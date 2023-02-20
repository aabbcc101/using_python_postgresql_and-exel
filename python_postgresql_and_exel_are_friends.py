# python, postgresql and excel are friends
# I wanna show you how to use it in pactice
# use parser to keep your credentials
#task find all blocked devices and fill a table with these data 


from openpyxl import load_workbook
import openpyxl
import psycopg2

from configparser import ConfigParser

def config(filename='C:\\Users\\vlad\\Documents\\DO_NOT_SHARE\\dbases.ini', section='postgresql_database_1'):
    # if u use Linux change the path to credentials 

    parser = ConfigParser()
    parser.read(filename)

    db = {}

    if parser.has_section(section):
        params = parser.items(section)
        for param in params:
            db[param[0]] = param[1]
    else:
        raise Exception('Section {0} not found in the {1} file'.format(section, filename))

    return db
    
wb = load_workbook('read_123.xlsx')  #exel with data: birthday, _, last_name, first_name , middle_name and so on
sheet = wb['Sheet1']

wb_write = openpyxl.Workbook()
wb_write.create_sheet(title = 'Первый лист', index = 0) # a litlle Russian would be nice. At least for a test
sheet_write = wb_write['Первый лист']

#sheet['A1'] = 'Serial number'  I don't need it, I just left it here for some options
sheet_write['B1'] = 'last name'
sheet_write['C1'] = 'first name'
sheet_write['D1'] = 'middle name'
sheet_write['E1'] = 'BITHDAY'
sheet_write['F1'] = 'Client_id'
sheet_write['G1'] = 'login'
sheet_write['H1'] = 'blocked'
sheet_write['I1'] = 'connected'

list_first_names = []
list_middle_names = []
list_last_names = []
birth_days = []
client_id_list = []
exel_row_number = 1

# getting data from the exel file

for birthday, _, last_name, first_name , middle_name in sheet['H2':'L156']: #['H81':'L82']: #['H2':'L156']:
    birth_days.append(birthday.value)
    list_last_names.append(last_name.value)
    list_first_names.append(first_name.value)
    list_middle_names.append(middle_name.value)
        
connect_postgresql = psycopg2.connect(**config())   
    
cur = connect_postgresql.cursor()

for number in range(len(list_last_names)):#[0,1]: testing part

    exel_row_number = exel_row_number + 1
    birth_days[number] = str(birth_days[number])[:10]

    #getting client_id
   
    select_client_id = ("select client_id from client_information "
        "where LOWER(last_name) LIKE LOWER('" + str(list_last_names[number]) + "%') "
        "and LOWER(first_name) LIKE LOWER('" + str(list_first_names[number]) + "%') "
        "and LOWER(middle_name) LIKE LOWER('" + str(list_middle_names[number]) + "%') "
        "and client_birth_date = '" + str(birth_days[number]) + "'; ")

    cur.execute(select_client_id)
    client_id = cur.fetchall()
    if client_id:
        print ("\n--------------------------------------------------------------\n")

        select_user_device = ("SELECT login, blocked, connected  FROM devices where client_id in ('" + str(client_id[0][0]) + "') order by create_date desc;")
            
        cur.execute(select_user_device)
        login_blocked_connected = cur.fetchall()
        
        # write last_name, first_name, middle_name, birthday, client_id down and print it
        print(str(exel_row_number) + " " + list_last_names[number] + " " + list_first_names[number] + " " + list_middle_names[number]+ " " + birth_days[number] + " ")
        cell = sheet_write.cell(row = exel_row_number, column = 2)
        cell.value = "".join(list_last_names[number])
        cell = sheet_write.cell(row = exel_row_number, column = 3)
        cell.value = "".join(list_first_names[number])
        cell = sheet_write.cell(row = exel_row_number, column = 4)
        cell.value = "".join(list_middle_names[number])
        cell = sheet_write.cell(row = exel_row_number, column = 5)
        cell.value = "".join(birth_days[number])
        cell = sheet_write.cell(row = exel_row_number, column = 6)
        cell.value = str(client_id[0][0])
        
        print_string = ''
        string_login = ''
        string_blocked = ''
        string_connected = ''
        
        # it can contains a lot of raws. This way is the best to read in the teble of Exel
        for login, blocked, connected in login_blocked_connected:  
            print_string += str(client_id[0][0]) + " " + str(login) + " " + str(blocked) + " " + str(connected) + '\n'
            string_login += str(login)+ '\n'
            string_blocked += str(blocked)+ '\n'
            string_connected += str(connected)+ '\n'
            print(print_string, end='')

        cell = sheet_write.cell(row = exel_row_number, column = 7)
        cell.value = "".join(string_login)
        cell = sheet_write.cell(row = exel_row_number, column = 8)
        cell.value = str(string_blocked)
        cell = sheet_write.cell(row = exel_row_number, column = 9)
        cell.value = str(string_connected)
    else:
        print ("\n--------------------------------------------------------------\n")    
        print(str(exel_row_number) + " " + list_last_names[number] + " " + list_first_names[number] + " " + list_middle_names[number]+ " " + birth_days[number] + " "  + " не найдено")
        
        cell = sheet_write.cell(row = exel_row_number, column = 2)
        cell.value = "".join(list_last_names[number])
        cell = sheet_write.cell(row = exel_row_number, column = 3)
        cell.value = "".join(list_first_names[number])
        cell = sheet_write.cell(row = exel_row_number, column = 4)
        cell.value = "".join(list_middle_names[number])
        cell = sheet_write.cell(row = exel_row_number, column = 5)
        cell.value = "".join(birth_days[number])
        cell = sheet_write.cell(row = exel_row_number, column = 6)
        cell.value = "нет данных"
        cell = sheet_write.cell(row = exel_row_number, column = 7)
        cell.value = "нет данных"
        cell = sheet_write.cell(row = exel_row_number, column = 8)
        cell.value = "нет данных"
        cell = sheet_write.cell(row = exel_row_number, column = 9)
        cell.value = "нет данных"        
  
    
wb_write.save('result_table.xlsx')       
        
cur.close()
connect_postgresql.close()    

